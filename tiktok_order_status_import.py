"""Parser สำหรับไฟล์ "ทั้งหมด คำสั่งซื้อ" export จาก TikTok Shop Seller Center
(คำสั่งซื้อ > ทั้งหมด > Export ทุกคอลัมน์ > ชีต "OrderSKUList") — ใช้แค่คอลัมน์
"Order Status" มายืนยันว่าออเดอร์ไหนถูกยกเลิกจริง (ไม่ได้ใช้ข้อมูลอื่นในไฟล์นี้เลย)

พบ 2026-09-05 ระหว่างช่วยผู้ใช้เช็คกำไร TikTok ว่า sync_tiktok_to_ecommerce()
เช็คการยกเลิกออเดอร์จาก order_status ของไฟล์ affiliate_orders/income เท่านั้น
("ไม่มีสิทธิ์"/"ชำระแล้ว"/"ลูกค้ายังไม่ได้ชำระเงิน" ฯลฯ — เป็นคำศัพท์ของรายงานนายหน้า/
การเงิน ไม่เคยมีค่า "ยกเลิกแล้ว" เป๊ะๆ เลย) ทำให้ filter "if order_status ==
'ยกเลิกแล้ว': skip" ใน ecom_calc.aggregate_order_costs ไม่เคยตัดออเดอร์ที่ยกเลิก
จริงออกเลยสักออเดอร์ — พัสดุ COD ที่ลูกค้าปฏิเสธ/ถูกตีกลับ ถูกนับเป็นขาดทุนเต็มต้นทุน
ทั้งที่ไม่ได้ขายจริง ไฟล์นี้ (จาก "คำสั่งซื้อ > ทั้งหมด") เป็นแหล่งเดียวที่บอกสถานะ
ยกเลิกตรงๆ ด้วยคำว่า "ยกเลิกแล้ว" เป๊ะ

หมายเหตุ pandas.read_excel() (ทั้ง engine="openpyxl" default และ explicit) อ่านไฟล์
นี้ผิดเหลือแค่คอลัมน์เดียว ("Order ID") ทั้งที่ไฟล์จริงมี 65 คอลัมน์ (ยืนยันด้วย
openpyxl.load_workbook() ตรงๆ ว่าอ่านครบ) — สาเหตุไม่ทราบแน่ชัด (ไม่ใช่ hidden
column/Table object/AutoFilter อย่างที่สงสัยตอนแรก) เลี่ยงปัญหาด้วยการอ่านผ่าน
openpyxl โดยตรงแทน pandas.read_excel ในไฟล์นี้ไฟล์เดียว"""
import openpyxl

from ecom_import_common import id_str_or_none as _id_str_or_none

_SHEET_NAME = "OrderSKUList"


def parse_order_statuses(file) -> dict[str, str]:
    """อ่านชีต 'OrderSKUList' คืน {order_id: order_status} (แถวที่ 1 = หัวคอลัมน์,
    แถวที่ 2 = คำอธิบายฟิลด์ ข้ามทิ้ง, ข้อมูลจริงเริ่มแถว 3) — ออเดอร์เดียวอาจมีหลาย
    แถว (หลาย SKU) แต่ทุกแถวของออเดอร์เดียวกันใช้สถานะเดียวกันเสมอ ใช้ค่าแถวแรกที่เจอ"""
    wb = openpyxl.load_workbook(file, data_only=True)
    if _SHEET_NAME not in wb.sheetnames:
        return {}
    ws = wb[_SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return {}
    header = rows[0]
    try:
        oid_idx = header.index("Order ID")
        status_idx = header.index("Order Status")
    except ValueError:
        return {}

    out: dict[str, str] = {}
    for r in rows[2:]:
        oid = _id_str_or_none(r[oid_idx])
        if not oid or oid in out:
            continue
        status = r[status_idx]
        if status:
            out[oid] = str(status)
    return out
